import pandas as pd
from selenium import webdriver
import openpyxl as op
import pyautogui
from time import sleep
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys


class ChromeAuto:
    def __init__(self):
        self.driver_path = 'chromedriver'
        self.options = webdriver.ChromeOptions()
        self.options.add_argument('--profile-directory=1')
        self.chrome = webdriver.Chrome(
            self.driver_path,
            options=self.options
        )

    def acessar(self, site):
        self.chrome.get(site)

    def login(self):
        self.chrome.maximize_window()
        logim = self.chrome.find_element_by_css_selector('#EUsuario_CAMPO')
        senha = self.chrome.find_element_by_css_selector('#ESenha_CAMPO')
        entrar = self.chrome.find_element_by_css_selector('#lnkEntrar')

        sleep(5)
        logim.send_keys('34138047824_885134')

        senha.send_keys('dIGITAL@21')

        entrar.click()
        sleep(5)

        auto= self.chrome.find_element_by_xpath('//*[@id="content"]/div/div/div[1]/div/div[1]').click()

        auto1= self.chrome.find_element_by_xpath('//*[@id="content"]/div/div/div[1]/div/div[1]/div/div[2]/div[3]').click()
        sleep(5)
        inss = self.chrome.find_element_by_xpath('//*[@id="navbar-collapse-funcao"]/ul/li[5]/a')
        sleep(10)

        inss.click()
        sleep(10)
        inss.click()
        sleep(10)
        auto2 = self.chrome.find_element_by_css_selector(
            '#navbar-collapse-funcao > ul > li.open > ul > li:nth-child(1)')
        auto2.click()
        import pyautogui
        self.chrome.refresh()

    def solicitacao(self):
        pedidos = op.load_workbook('fé.xlsx')
        nomes_planilhas = pedidos.sheetnames  # aqui vc ta pegadno quantas paginas tem no excel
        planilhas1 = pedidos['Worksheet']  # Aqui vc ta pegando tudo que tem na planilha

        dados = []
        for camp in planilhas1['a']:  # aqui vc ta pegando exatamente oq tem na coluna b da planilha
            if camp.value is not None:
                dados.append(camp.value)

        for index in range(len(dados)):
            try:
                dados1 = dados[index]
                sleep(5)
                sob = self.chrome.find_element_by_xpath(
                    '//*[@id="ctl00_Cph_jp1_pnlDadosBeneficiario_Container_AbaTermoAutorizacao_rdlTipoTermo_1"]').click()
                voltar = self.chrome.find_element_by_xpath(
                    '//*[@id="__tab_ctl00_Cph_jp1_pnlDadosBeneficiario_Container_AbaTermoAutorizacao"]').click()
                sleep(4)

                cpf = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_Container_AbaTermoAutorizacao_txtCpfCli_CAMPO').send_keys(dados1)
                pyautogui.click(x=85, y=175)


                sleep(1)
                try:
                    sleep(2)
                    alert = self.chrome.switch_to.alert
                    alert.accept()
                    sleep(1)
                except Exception:
                    print('sem mensagem')



                sleep(5)
                nome = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_Container_AbaTermoAutorizacao_txtNomeCli_CAMPO').send_keys('Lindalva pereira')
                sleep(1)
                cidade = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_Container_AbaTermoAutorizacao_txtLocalAssTermo_CAMPO').send_keys('São paulo')
                sleep(2)
                btn = self.chrome.find_element_by_xpath(
                    '//*[@id="ctl00_Cph_jp1_pnlDadosBeneficiario_Container_AbaTermoAutorizacao_btnImprimirTermo_dvCBtn"]').click()
                sleep(5)
                soli = self.chrome.find_element_by_css_selector(
                    '#__tab_ctl00_Cph_jp1_pnlDadosBeneficiario_Container_ConsultaDadosBeneficio')
                soli.click()

                sleep(3)

                chave = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_Container_ConsultaDadosBeneficio_cboChaveTermo_CAMPO').click()
                sleep(1)
                chave1= self.chrome.find_element_by_xpath('//*[@id="ctl00_Cph_jp1_pnlDadosBeneficiario_Container_ConsultaDadosBeneficio_cboChaveTermo_CAMPO"]/option[2]').click()
                sleep(3)

                tp = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_Container_ConsultaDadosBeneficio_cboDocIdentifCli_CAMPO')
                sleep(3)
                tp.click()
                sleep(3)
                rg=self.chrome.find_element_by_xpath('//*[@id="ctl00_Cph_jp1_pnlDadosBeneficiario_Container_ConsultaDadosBeneficio_cboDocIdentifCli_CAMPO"]/option[7]').click()
                sleep(4)
                anexo = self.chrome.find_element_by_xpath('//*[@id="ctl00_Cph_jp1_pnlDadosBeneficiario_Container_ConsultaDadosBeneficio_grdArquivosUpload_ctl02_btnAnexarArquivoUpload"]').click()


                sleep(3)

                pyautogui.click(x=83, y=177)
                sleep(1)
                pyautogui.click(x=83, y=177)
                sleep(1)
                pyautogui.click(x=283, y=172)
                pyautogui.click(x=283, y=172)
                sleep(1)
                up= self.chrome.find_element_by_xpath('//*[@id="ctl00_Cph_jp1_pnlDadosBeneficiario_Container_ConsultaDadosBeneficio_btnRealizarUpload_dvTxt"]/table').click()
                sleep(20)
                try:
                    anexo2 = self.chrome.find_element_by_xpath('//*[@id="ctl00_Cph_jp1_pnlDadosBeneficiario_Container_ConsultaDadosBeneficio_grdTermoAutBenef_ctl02_UpdatePanelEnviarArquivoGrid2"]/label').click()

                    sleep(3)

                    pyautogui.click(x=83, y=177)
                    sleep(1)
                    pyautogui.click(x=83, y=177)
                    sleep(1)
                    pyautogui.click(x=283, y=172)
                    pyautogui.click(x=283, y=172)
                    sleep(1)

                    btnau = self.chrome.find_element_by_css_selector('#ctl00_Cph_jp1_pnlDadosBeneficiario_Container_ConsultaDadosBeneficio_btnSolicitarAutorizacao_dvTxt > table > tbody > tr > td')
                    btnau.click()
                    sleep(20)

                except Exception :
                    print(' Teve mensagem')

                    alert = self.chrome.switch_to.alert
                    alert.accept()
                    sleep(1)
                    sleep(7)
                    anexo2 = self.chrome.find_element_by_xpath(
                        '//*[@id="ctl00_Cph_jp1_pnlDadosBeneficiario_Container_ConsultaDadosBeneficio_grdTermoAutBenef_ctl02_UpdatePanelEnviarArquivoGrid2"]/label').click()

                    sleep(3)

                    pyautogui.click(x=83, y=177)
                    sleep(1)
                    pyautogui.click(x=83, y=177)
                    sleep(1)
                    pyautogui.click(x=283, y=172)
                    pyautogui.click(x=283, y=172)
                    sleep(6)

                btnau = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_Container_ConsultaDadosBeneficio_btnSolicitarAutorizacao_dvTxt > table > tbody > tr > td')
                btnau.click()
                sleep(17)



                voltar = self.chrome.find_element_by_xpath(
                    '//*[@id="ctl00_Cph_jp1_pnlDadosBeneficiario_Container_AbaTermoAutorizacao_tab"]/span/span').click()

            except Exception as e:
                print(e)
                try:
                    voltar= self.chrome.find_element_by_xpath('//*[@id="ctl00_Cph_jp1_pnlDadosBeneficiario_Container_AbaTermoAutorizacao_tab"]/span/span').click()
                    sleep(1)
                    print('erro')
                except Exception:
                    self.chrome.refresh()
                    voltar = self.chrome.find_element_by_xpath(
                        '//*[@id="ctl00_Cph_jp1_pnlDadosBeneficiario_Container_AbaTermoAutorizacao_tab"]/span/span').click()
                    sleep(1)
                    print('erro')



    def ponte(self):
        chrome.acessar('https://consignado.bancopaulista.com.br/WebAutorizador/?FISession=b63d5dd7e5f1')
        inss = self.chrome.find_element_by_css_selector('#navbar-collapse-funcao > ul > li:nth-child(5) > a')
        inss.click()
        sleep(5)
        inss.click()
        sleep(5)
        auto = self.chrome.find_element_by_xpath('//*[@id="WFP2010_MTCNSBENEF"]')
        auto.click()


    def consulta(self):

        pedidos = op.load_workbook('fé.xlsx')
        nomes_planilhas = pedidos.sheetnames  # aqui vc ta pegadno quantas paginas tem no excel
        planilhas1 = pedidos['Worksheet']  # Aqui vc ta pegando tudo que tem na planilha

        dados = []
        for camp in planilhas1['a']:  # aqui vc ta pegando exatamente oq tem na coluna b da planilha
            if camp.value is not None:
                dados.append(camp.value)
        attivoL = []
        beneL = []
        dataaL = []
        ufL = []
        especieL = []
        tipodcontaL = []
        cbcL = []
        agenciaL = []
        conta_correnteL = []
        classifiL = []
        legalL = []
        procuradorL = []
        edidadeL = []
        emprestimoL = []
        despachoL = []
        margemL = []
        cartL = []
        limiteL = []
        ativosL = []
        data_consultaL = []
        representadoL = []
        nome_represenL = []
        elegivel_emprestimoL = []
        for index in range(len(dados)):
            try:

                sleep(10)
                cliente = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_txtCPFCliente_CAMPO')
                dados1 = dados[index]
                sleep(3)

                cliente.send_keys(dados1)
                sleep(4)
                pyautogui.press('enter')
                try:
                    sleep(2)
                    alert = self.chrome.switch_to.alert
                    alert.accept()
                    sleep(1)
                except Exception:
                    print('sem mensagem')
                sleep(4)
                nome_clien = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_ddlNomeCli_CAMPO')
                nome_clien.click()
                pyautogui.press('down')
                pyautogui.press('enter')
                sleep(4)

                btn = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_btnObterBeneficios_dvTxt > table > tbody > tr > td')
                btn.click()
                btn.click()
                sleep(5)

                validos = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_ddlBeneficios_CAMPO')
                validos.click()
                pyautogui.press('down')
                pyautogui.press('enter')
                sleep(2)
                btn1 = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_btnConsultarBeneficio_dvTxt > table > tbody > tr > td')
                btn1.click()
                sleep(4)
                attivo = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(5) > td:nth-child(2)').text
                bene = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(7) > td:nth-child(2)').text
                especie = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(6) > td:nth-child(2)').text
                dataa = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(8) > td:nth-child(2)').text
                uf = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(9) > td:nth-child(2)').text
                tipodconta = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(10) > td:nth-child(2)').text
                cbc = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(11) > td:nth-child(2)').text
                agencia = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(12) > td:nth-child(2)').text
                conta_corrente = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(13) > td:nth-child(2)').text
                classifi = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(14) > td:nth-child(2)').text
                legal = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(15) > td:nth-child(2)').text
                procurador = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(16) > td:nth-child(2)').text
                edidade = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(17) > td:nth-child(2)').text
                emprestimo = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(18) > td:nth-child(2)').text
                despacho = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(20) > td:nth-child(2)').text
                margem = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(21) > td:nth-child(2)').text
                cart = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(22) > td:nth-child(2)').text
                limite = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(23) > td:nth-child(2)').text
                ativos = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(24) > td:nth-child(2)').text
                data_consulta = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(25) > td:nth-child(2)').text
                representado = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(26) > td:nth-child(2)').text
                nome_represen = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(27) > td:nth-child(2)').text
                elegivel_emprestimo = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_lblRetorno > table > tbody > tr > td > table > tbody > tr:nth-child(29) > td:nth-child(2)').text
                sleep(2)
                attivoL.append(attivo)
                beneL.append(bene)
                dataaL.append(dataa)
                ufL.append(uf)
                especieL.append(especie)
                tipodcontaL.append(tipodconta)
                cbcL.append(cbc)
                agenciaL.append(agencia)
                conta_correnteL.append(conta_corrente)
                classifiL.append(classifi)
                legalL.append(legal)
                procuradorL.append(procurador)
                edidadeL.append(edidade)
                emprestimoL.append(emprestimo)
                despachoL.append(despacho)
                margemL.append(margem)
                cartL.append(cart)
                limiteL.append(limite)
                ativosL.append(ativos)
                data_consultaL.append(data_consulta)
                representadoL.append(representado)
                nome_represenL.append(nome_represen)
                elegivel_emprestimoL.append(elegivel_emprestimo)
                sleep(2)
                cpfcliente = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_txtCPFCliente_CAMPO')
                cpfcliente.clear()


            except Exception as eee:
                cpfcliente = self.chrome.find_element_by_css_selector(
                    '#ctl00_Cph_jp1_pnlDadosBeneficiario_txtCPFCliente_CAMPO')
                cpfcliente.clear()

                pyautogui.press('enter')
                sleep(1)
                attivoL.append('Sem item')
                beneL.append('Sem item')
                dataaL.append('Sem item')
                ufL.append('Sem item')
                especieL.append('sem item')
                tipodcontaL.append('Sem item')
                cbcL.append('Sem item')
                agenciaL.append('Sem item')
                conta_correnteL.append('Sem item')
                classifiL.append('Sem item')
                legalL.append('Sem item')
                procuradorL.append('Sem item')
                edidadeL.append('Sem item')
                emprestimoL.append('Sem item')
                despachoL.append('Sem item')
                margemL.append('Sem item')
                cartL.append('Sem item')
                limiteL.append('Sem item')
                ativosL.append('Sem item')
                data_consultaL.append('Sem item')
                representadoL.append('Sem item')
                nome_represenL.append('Sem item')
                elegivel_emprestimoL.append('Sem item')
                sleep(2)

        data1 = dados
        df2 = pd.DataFrame(data1)
        df1 = df2.assign(Situação=attivoL,
                         Benefício_concedido_por_liminar=beneL,
                         Data_de_Cessação_do_Benefício=dataaL,
                         UF=ufL,
                         Espécie_do_Benefício=especieL,
                         Tipodeconta=tipodcontaL,
                         Cbl=cbcL,
                         Agencia=agenciaL,
                         Conta_corrente=conta_correnteL,
                         Classificador_da_Pensãoa_limentícia=classifiL,
                         Possui_representante_legal=legalL,
                         Possui_procurador=procuradorL,
                         Possui_entidade=edidadeL,
                         Benefício_bloqueado_para_empréstimo=emprestimoL,
                         Data_do_Despacho_do_Benefício=despachoL,
                         Margem=margemL,
                         Margem_de_cartão=cartL,
                         limitedocartão=limiteL,
                         Situacao=attivoL,
                         data_dac_onsulta=data_consultaL,
                         Representa=representadoL,
                         nome_do_representando=nome_represenL,
                         elegivel_empretimo=elegivel_emprestimoL)
        print(df1)
        df1.to_excel('Rafpronta706.xlsx')


if __name__ == '__main__':
    chrome = ChromeAuto()
    chrome.acessar(
        'https://consignado.bancopaulista.com.br')
    chrome.login()
    sleep(3)
    chrome.solicitacao()
    sleep(1)
    chrome.ponte()
    sleep(2)
    chrome.consulta()